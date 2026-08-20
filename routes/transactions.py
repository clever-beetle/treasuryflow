from flask import Blueprint, render_template, request, url_for, redirect, session, flash, Response, jsonify, send_file, make_response
from utils import get_db, login_required, CATEGORIES
from datetime import datetime
import io
import csv
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

transactions_bp = Blueprint('transactions', __name__)

@transactions_bp.route('/calendar', strict_slashes=False, endpoint='transactions_calendar')
@login_required
def transactions_calendar():
    from flask import request
    # Pass all query args to transactions_list but override view
    args = request.args.copy()
    args['view'] = 'calendar'
    request.args = args
    return transactions_list()

@transactions_bp.route('', strict_slashes=False)
@transactions_bp.route('/', strict_slashes=False)
@login_required
def transactions_list():
    db = get_db()
    user_id = session['user_id']
    view = request.args.get('view', 'list')
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    selected_category = request.args.get('category', '')
    selected_account = request.args.get('account_id', '')
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')
    
    # Base query
    query = '''
        SELECT t.*, a.name as account_name 
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        WHERE t.user_id = ?
    '''
    params = [user_id]
    
    if search_query:
        query += ' AND t.description LIKE ?'
        params.append(f'%{search_query}%')
    if start_date:
        query += ' AND t.date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND t.date <= ?'
        params.append(end_date)
    if selected_category:
        query += ' AND t.category = ?'
        params.append(selected_category)
    if selected_account:
        query += ' AND t.account_id = ?'
        params.append(selected_account)
        
    # Validate sort parameters to prevent SQL injection
    allowed_sort_columns = {'date', 'amount', 'category', 'description', 'type'}
    allowed_sort_orders = {'asc', 'desc'}
    if sort_by not in allowed_sort_columns:
        sort_by = 'date'
    if sort_order.lower() not in allowed_sort_orders:
        sort_order = 'desc'
        
    query += f' ORDER BY t.{sort_by} {sort_order}'
    
    transactions = db.execute(query, params).fetchall()
    
    # Calculate totals
    total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in transactions if t['type'] == 'expense')
    net_flow = total_income - total_expense
    
    # Pagination for list view
    per_page = 20
    total_tx = len(transactions)
    total_pages = (total_tx + per_page - 1) // per_page
    start = (page - 1) * per_page
    end = start + per_page
    
    tx_page = transactions[start:end]
    
    accounts = db.execute('SELECT * FROM accounts WHERE user_id = ?', (user_id,)).fetchall()
    
    # Calendar data
    import calendar
    now = datetime.now()
    now_date = now.date()
    cal_year = request.args.get('cal_year', now.year, type=int)
    cal_month = request.args.get('cal_month', now.month, type=int)
    
    cal = calendar.Calendar()
    month_days = cal.monthdatescalendar(cal_year, cal_month)
    
    from datetime import date as dt_date, datetime as dt_datetime
    cal_tx_map = {}
    for tx in transactions:
        d_val = tx['date']
        if isinstance(d_val, (dt_datetime, dt_date)):
            d_str = d_val.strftime('%Y-%m-%d')
        else:
            d_str = str(d_val)[:10]
        if d_str not in cal_tx_map:
            cal_tx_map[d_str] = {'income': 0, 'expense': 0}
        cal_tx_map[d_str][tx['type']] += tx['amount']
    
    return render_template('transactions_list.html', 
                           month_days=month_days,
                           transactions=tx_page, 
                           all_transactions=transactions,
                           total_income=total_income,
                           total_expense=total_expense,
                           net_flow=net_flow,
                           view=view,
                           page=page,
                           total_pages=total_pages,
                           search_query=search_query,
                           start_date=start_date,
                           end_date=end_date,
                           selected_category=selected_category,
                           selected_account=selected_account,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           accounts=accounts,
                           categories=CATEGORIES,
                           cal_year=cal_year,
                           cal_month=cal_month,
                           now=now_date,
                           cal_tx_map=cal_tx_map,
                           month_name=datetime(cal_year, cal_month, 1).strftime('%B'))

@transactions_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_transaction():
    db = get_db()
    user_id = session['user_id']
    if request.method == 'POST':
        t_type = request.form.get('type')
        date = request.form.get('date')
        amount_str = request.form.get('amount', '0').replace('.', '').replace(',', '')
        amount = float(amount_str) if amount_str else 0
        category = request.form.get('category')
        account_id = request.form.get('account_id')
        description = request.form.get('description', '')
        to_account_id = request.form.get('to_account_id')
        
        if t_type == 'transfer':
            if account_id == to_account_id:
                flash('Cannot transfer to same account', 'danger')
                return redirect(url_for('transactions.add_transaction'))
                
            db.execute('''
                INSERT INTO transactions (user_id, account_id, date, type, amount, description, category)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, account_id, date, 'expense', amount, f'Transfer to account {to_account_id}: {description}', 'Transfer'))
            
            db.execute('''
                INSERT INTO transactions (user_id, account_id, date, type, amount, description, category)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, to_account_id, date, 'income', amount, f'Transfer from account {account_id}: {description}', 'Transfer'))
            
            db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (amount, account_id))
            db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (amount, to_account_id))
        else:
            db.execute('''
                INSERT INTO transactions (user_id, account_id, date, type, amount, description, category)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, account_id, date, t_type, amount, description, category))
            
            if t_type == 'income':
                db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (amount, account_id))
            else:
                db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (amount, account_id))
                
        db.commit()
        flash('Transaction added', 'success')
        return redirect(url_for('transactions.transactions_list'))
        
    accounts = db.execute('SELECT * FROM accounts WHERE user_id = ?', (user_id,)).fetchall()
    return render_template('add_transaction.html', accounts=accounts, categories=CATEGORIES, now=datetime.now())

@transactions_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_transaction(id):
    db = get_db()
    user_id = session['user_id']
    tx = db.execute('SELECT * FROM transactions WHERE id = ? AND user_id = ?', (id, user_id)).fetchone()
    if not tx:
        flash('Transaction not found', 'danger')
        return redirect(url_for('transactions.transactions_list'))
        
    if request.method == 'POST':
        # Revert old balance
        if tx['type'] == 'income':
            db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (tx['amount'], tx['account_id']))
        else:
            db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (tx['amount'], tx['account_id']))
            
        # Apply new
        t_type = request.form.get('type')
        date = request.form.get('date')
        amount_str = request.form.get('amount', '0').replace('.', '').replace(',', '')
        amount = float(amount_str) if amount_str else 0
        category = request.form.get('category')
        account_id = request.form.get('account_id')
        description = request.form.get('description', '')
        
        db.execute('''
            UPDATE transactions 
            SET account_id=?, date=?, type=?, amount=?, description=?, category=?
            WHERE id=? AND user_id=?
        ''', (account_id, date, t_type, amount, description, category, id, user_id))
        
        if t_type == 'income':
            db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (amount, account_id))
        else:
            db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (amount, account_id))
            
        db.commit()
        flash('Transaction updated', 'success')
        return redirect(url_for('transactions.transactions_list'))
        
    accounts = db.execute('SELECT * FROM accounts WHERE user_id = ?', (user_id,)).fetchall()
    return render_template('add_transaction.html', tx=tx, accounts=accounts, categories=CATEGORIES)

@transactions_bp.route('/delete/<int:transaction_id>', methods=['POST'])
@login_required
def delete_transaction(transaction_id):
    db = get_db()
    user_id = session['user_id']
    tx = db.execute('SELECT * FROM transactions WHERE id = ? AND user_id = ?', (transaction_id, user_id)).fetchone()
    if tx:
        if tx['type'] == 'income':
            db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (tx['amount'], tx['account_id']))
        else:
            db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (tx['amount'], tx['account_id']))
        db.execute('DELETE FROM transactions WHERE id = ?', (transaction_id,))
        db.commit()
        flash('Transaction deleted', 'success')
    return redirect(url_for('transactions.transactions_list'))

@transactions_bp.route('/bulk_delete', methods=['POST'])
@login_required
def bulk_delete_transactions():
    db = get_db()
    user_id = session['user_id']
    tx_ids = request.form.getlist('transaction_ids')
    for tid in tx_ids:
        tx = db.execute('SELECT * FROM transactions WHERE id = ? AND user_id = ?', (tid, user_id)).fetchone()
        if tx:
            if tx['type'] == 'income':
                db.execute('UPDATE accounts SET current_balance = current_balance - ? WHERE id = ?', (tx['amount'], tx['account_id']))
            else:
                db.execute('UPDATE accounts SET current_balance = current_balance + ? WHERE id = ?', (tx['amount'], tx['account_id']))
            db.execute('DELETE FROM transactions WHERE id = ?', (tid,))
    db.commit()
    flash('Selected transactions deleted', 'success')
    return redirect(url_for('transactions.transactions_list'))

@transactions_bp.route('/bulk_edit', methods=['POST'])
@login_required
def bulk_edit_transactions():
    db = get_db()
    user_id = session['user_id']
    tx_ids = request.form.get('transaction_ids_bulk', '').split(',')
    category = request.form.get('bulk_category')
    
    if category and tx_ids:
        for tid in tx_ids:
            if tid:
                db.execute('UPDATE transactions SET category = ? WHERE id = ? AND user_id = ?', (category, tid, user_id))
        db.commit()
        flash('Transactions updated', 'success')
    return redirect(url_for('transactions.transactions_list'))

@transactions_bp.route('/export_csv')
@login_required
def export_csv():
    db = get_db()
    user_id = session['user_id']
    transactions = db.execute('''
        SELECT t.date, t.type, t.amount, t.description, t.category, a.name as account_name 
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        WHERE t.user_id = ? ORDER BY t.date DESC
    ''', (user_id,)).fetchall()
    
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(['Date', 'Type', 'Amount', 'Description', 'Category', 'Account'])
    for t in transactions:
        cw.writerow([t['date'], t['type'], t['amount'], t['description'], t['category'], t['account_name']])
        
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=transactions.csv"
    output.headers["Content-type"] = "text/csv"
    return output

@transactions_bp.route('/export_excel')
@login_required
def export_excel():
    db = get_db()
    user_id = session['user_id']
    transactions = db.execute('''
        SELECT t.date, t.type, t.amount, t.description, t.category, a.name as account_name 
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        WHERE t.user_id = ? ORDER BY t.date DESC
    ''', (user_id,)).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ['Date', 'Type', 'Amount (Rp)', 'Description', 'Category', 'Account']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="11237E", end_color="11237E", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    cat_totals = {}
    row_idx = 2
    for tx in transactions:
        amount = tx['amount']
        type_cap = tx['type'].capitalize() if tx['type'] else ''
        ws.append([tx['date'], type_cap, amount, tx['description'], tx['category'], tx['account_name']])
        if tx['type'] == 'expense':
            cat = tx['category']
            cat_totals[cat] = cat_totals.get(cat, 0) + amount
        ws.cell(row=row_idx, column=3).number_format = '#,##0'
        row_idx += 1

    if cat_totals:
        ws_chart = wb.create_sheet("Expense Analysis")
        ws_chart.append(["Category", "Total Spent"])
        r = 2
        for cat, total in cat_totals.items():
            ws_chart.append([cat, total])
            r += 1
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Expenses by Category"
        chart.y_axis.title = 'Amount (Rp)'
        chart.x_axis.title = 'Category'
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=r-1, max_col=2)
        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=r-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_chart.add_chart(chart, "D2")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="finance_report.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@transactions_bp.route('/export_pdf')
@login_required
def export_pdf():
    db = get_db()
    user_id = session['user_id']
    user = db.execute('SELECT fullname, username FROM users WHERE id = ?', (user_id,)).fetchone()
    fullname = user['fullname'] if user and user['fullname'] else (user['username'] if user else 'User')
    
    # Get filter params for period display
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = '''
        SELECT t.date, t.type, t.amount, t.description, t.category, a.name as account_name 
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        WHERE t.user_id = ?
    '''
    params = [user_id]
    if start_date:
        query += ' AND t.date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND t.date <= ?'
        params.append(end_date)
    query += ' ORDER BY t.date DESC'
    
    transactions = db.execute(query, params).fetchall()
    
    total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
    total_expense = sum(t['amount'] for t in transactions if t['type'] == 'expense')
    net_flow = total_income - total_expense
    total_transactions = len(transactions)

    # --- Period label ---
    if start_date and end_date:
        period_label = f"{start_date}  to  {end_date}"
    elif start_date:
        period_label = f"{start_date}  to  Present"
    elif end_date:
        period_label = f"All Time  to  {end_date}"
    else:
        period_label = "All Time"

    # --- Custom PDF class ---
    class TreasuryPDF(FPDF):
        NAVY = (17, 35, 126)        # #11237E
        DARK_NAVY = (10, 20, 80)
        GOLD = (180, 150, 60)
        LIGHT_GRAY = (245, 245, 248)
        MID_GRAY = (200, 200, 210)
        DARK_GRAY = (100, 100, 110)
        WHITE = (255, 255, 255)
        GREEN = (16, 150, 72)
        RED = (200, 40, 40)
        BLACK = (30, 30, 30)
        ACCENT_BLUE = (60, 100, 200)
        logo_path = None

        def header(self):
            # Top accent bar
            self.set_fill_color(*self.NAVY)
            self.rect(0, 0, 210, 3, 'F')
            # Gold thin line below
            self.set_fill_color(*self.GOLD)
            self.rect(0, 3, 210, 0.7, 'F')

            # Logo on the left
            if self.logo_path:
                try:
                    self.image(self.logo_path, 14, 6, 22)
                except Exception:
                    pass

            # Company name (shifted right to balance with logo)
            self.set_y(7)
            self.set_font('Arial', 'B', 22)
            self.set_text_color(*self.NAVY)
            self.cell(0, 10, 'TREASURY FLOW', 0, 1, 'C')

            # Subtitle
            self.set_font('Arial', '', 9)
            self.set_text_color(*self.DARK_GRAY)
            self.cell(0, 5, 'Personal Wealth Management & Financial Analytics', 0, 1, 'C')
            
            # Contact info
            self.set_font('Arial', '', 7.5)
            self.set_text_color(*self.ACCENT_BLUE)
            self.cell(0, 4, 'treasuryflow@gmail.com  |  www.treasuryflow.web.id', 0, 1, 'C')

            # Header bottom border
            self.set_y(30)
            self.set_fill_color(*self.NAVY)
            self.rect(10, 30, 190, 0.5, 'F')
            self.set_fill_color(*self.GOLD)
            self.rect(10, 30.5, 190, 0.3, 'F')
            self.ln(6)

        def footer(self):
            self.set_y(-25)
            # Top line
            self.set_fill_color(*self.MID_GRAY)
            self.rect(10, self.get_y(), 190, 0.3, 'F')
            self.ln(3)

            # Powered by
            self.set_font('Arial', 'B', 7)
            self.set_text_color(*self.NAVY)
            self.cell(0, 4, 'Powered by Fernando Capital', 0, 1, 'C')

            # Confidential notice
            self.set_font('Arial', 'I', 6)
            self.set_text_color(*self.DARK_GRAY)
            self.cell(0, 3, 'This document is confidential and intended solely for the named recipient.', 0, 1, 'C')

            # Page number
            self.set_font('Arial', '', 7)
            self.set_text_color(*self.ACCENT_BLUE)
            self.cell(0, 4, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')

        def section_title(self, title):
            self.set_font('Arial', 'B', 12)
            self.set_text_color(*self.NAVY)
            self.cell(0, 8, title, 0, 1, 'L')
            # Underline accent
            self.set_fill_color(*self.GOLD)
            self.rect(self.get_x(), self.get_y(), 40, 0.6, 'F')
            self.ln(3)

        def summary_card(self, x, label, value, color):
            card_w = 58
            card_h = 22
            # Card background
            self.set_fill_color(*self.LIGHT_GRAY)
            self.rounded_rect(x, self.get_y(), card_w, card_h, 2, 'F')
            # Border top accent
            self.set_fill_color(*color)
            self.rect(x, self.get_y(), card_w, 1.2, 'F')
            # Label
            self.set_xy(x + 3, self.get_y() + 3)
            self.set_font('Arial', '', 6.5)
            self.set_text_color(*self.DARK_GRAY)
            self.cell(card_w - 6, 4, label, 0, 0, 'L')
            # Value
            self.set_xy(x + 3, self.get_y() + 5)
            self.set_font('Arial', 'B', 11)
            self.set_text_color(*color)
            self.cell(card_w - 6, 6, value, 0, 0, 'L')

        def rounded_rect(self, x, y, w, h, r, style=''):
            # Simple filled rect (FPDF doesn't support rounded natively)
            self.rect(x, y, w, h, style)

    # --- Build PDF ---
    pdf = TreasuryPDF()
    # Set logo path
    import os
    from flask import current_app
    logo_file = os.path.join(current_app.static_folder, 'img', 'logo_v4.png')
    if os.path.exists(logo_file):
        pdf.logo_path = logo_file
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.add_page()

    # === DOCUMENT TITLE ===
    pdf.set_font('Arial', 'B', 16)
    pdf.set_text_color(*TreasuryPDF.NAVY)
    pdf.cell(0, 10, 'FINANCIAL STATEMENT REPORT', 0, 1, 'C')
    pdf.ln(2)

    # === REPORT INFO BOX ===
    box_y = pdf.get_y()
    pdf.set_fill_color(*TreasuryPDF.LIGHT_GRAY)
    pdf.rect(10, box_y, 190, 20, 'F')
    # Left accent
    pdf.set_fill_color(*TreasuryPDF.NAVY)
    pdf.rect(10, box_y, 1.5, 20, 'F')
    # Left side
    pdf.set_xy(16, box_y + 3)
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
    pdf.cell(25, 5, 'Prepared For', 0, 0)
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(*TreasuryPDF.BLACK)
    pdf.cell(60, 5, f':  {fullname}', 0, 1)
    pdf.set_x(16)
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
    pdf.cell(25, 5, 'Period', 0, 0)
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(*TreasuryPDF.BLACK)
    pdf.cell(60, 5, f':  {period_label}', 0, 0)
    # Right side
    pdf.set_xy(130, box_y + 3)
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
    pdf.cell(25, 5, 'Report Date', 0, 0)
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(*TreasuryPDF.BLACK)
    report_date = datetime.now().strftime('%d %B %Y')
    pdf.cell(40, 5, f':  {report_date}', 0, 1)
    pdf.set_xy(130, box_y + 9)
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
    pdf.cell(25, 5, 'Total Entries', 0, 0)
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(*TreasuryPDF.BLACK)
    pdf.cell(40, 5, f':  {total_transactions}', 0, 1)

    pdf.set_y(box_y + 23)

    # === FINANCIAL SUMMARY ===
    pdf.section_title('Financial Summary')
    
    card_y = pdf.get_y()
    pdf.summary_card(10, 'TOTAL CASH IN (INCOME)', f'+ Rp {total_income:,.0f}', TreasuryPDF.GREEN)
    pdf.set_y(card_y)
    pdf.summary_card(72, 'TOTAL CASH OUT (EXPENSE)', f'- Rp {total_expense:,.0f}', TreasuryPDF.RED)
    pdf.set_y(card_y)
    net_color = TreasuryPDF.GREEN if net_flow >= 0 else TreasuryPDF.RED
    net_sign = '+' if net_flow >= 0 else '-'
    pdf.summary_card(134, 'NET FLOW / SAVINGS', f'{net_sign} Rp {abs(net_flow):,.0f}', net_color)

    pdf.set_y(card_y + 27)

    # === EXECUTIVE SUMMARY ===
    pdf.set_fill_color(235, 240, 255)
    exec_y = pdf.get_y()
    pdf.rect(10, exec_y, 190, 16, 'F')
    pdf.set_fill_color(*TreasuryPDF.ACCENT_BLUE)
    pdf.rect(10, exec_y, 1.5, 16, 'F')
    pdf.set_xy(16, exec_y + 2)
    pdf.set_font('Arial', 'B', 8)
    pdf.set_text_color(*TreasuryPDF.NAVY)
    pdf.cell(0, 4, 'Executive Summary', 0, 1)
    pdf.set_x(16)
    pdf.set_font('Arial', '', 7.5)
    pdf.set_text_color(*TreasuryPDF.BLACK)
    if total_income > 0:
        savings_pct = (net_flow / total_income) * 100
        if net_flow >= 0:
            summary_text = f'During this period, you earned Rp {total_income:,.0f} and spent Rp {total_expense:,.0f}, resulting in a surplus of Rp {net_flow:,.0f} ({savings_pct:.1f}% savings rate). Keep it up!'
        else:
            summary_text = f'During this period, you earned Rp {total_income:,.0f} and spent Rp {total_expense:,.0f}, resulting in a deficit of Rp {abs(net_flow):,.0f}. Consider reducing expenses.'
    elif total_expense > 0:
        summary_text = f'During this period, you recorded Rp {total_expense:,.0f} in expenses with no income recorded. Please ensure all income transactions are logged.'
    else:
        summary_text = 'No transactions recorded for this period.'
    pdf.multi_cell(180, 3.5, summary_text, 0, 'L')

    pdf.set_y(exec_y + 19)

    # === TOP EXPENSE CATEGORIES ===
    cat_map = {}
    for t in transactions:
        if t['type'] == 'expense' and t['category'] and t['category'] != 'Transfer':
            cat_map[t['category']] = cat_map.get(t['category'], 0) + (t['amount'] or 0)
    
    if cat_map:
        pdf.section_title('Top Expense Categories')
        sorted_cats = sorted(cat_map.items(), key=lambda x: x[1], reverse=True)[:5]
        
        # Category bar chart
        bar_colors = [
            (17, 35, 126),    # Navy
            (60, 100, 200),   # Blue
            (180, 150, 60),   # Gold
            (200, 40, 40),    # Red
            (100, 100, 110),  # Gray
        ]
        max_val = sorted_cats[0][1] if sorted_cats else 1
        
        for ci, (cat_name, cat_total) in enumerate(sorted_cats):
            row_y = pdf.get_y()
            pct = (cat_total / total_expense * 100) if total_expense > 0 else 0
            bar_width = (cat_total / max_val) * 100
            
            # Category name
            pdf.set_font('Arial', '', 7)
            pdf.set_text_color(*TreasuryPDF.BLACK)
            pdf.cell(38, 5, f'{ci+1}. {cat_name[:20]}', 0, 0, 'L')
            
            # Bar background
            pdf.set_fill_color(235, 235, 240)
            pdf.rect(pdf.get_x(), row_y + 0.5, 100, 4, 'F')
            # Bar fill
            color = bar_colors[ci % len(bar_colors)]
            pdf.set_fill_color(*color)
            pdf.rect(pdf.get_x(), row_y + 0.5, bar_width, 4, 'F')
            pdf.cell(102, 5, '', 0, 0)
            
            # Amount & percentage
            pdf.set_font('Arial', 'B', 7)
            pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
            pdf.cell(50, 5, f'Rp {cat_total:,.0f}  ({pct:.1f}%)', 0, 1, 'R')
        
        pdf.ln(3)

    # === TRANSACTION TABLE ===
    pdf.section_title('Transaction Ledger')

    # Table header
    col_widths = [10, 20, 48, 32, 35, 25, 20]
    headers = ['NO.', 'DATE', 'DESCRIPTION', 'CATEGORY', 'ACCOUNT', 'AMOUNT', 'TYPE']
    
    def draw_table_header():
        pdf.set_fill_color(*TreasuryPDF.NAVY)
        pdf.set_text_color(*TreasuryPDF.WHITE)
        pdf.set_font('Arial', 'B', 6.5)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, 0, 0, 'C', fill=True)
        pdf.ln()
        # Gold line under header
        pdf.set_fill_color(*TreasuryPDF.GOLD)
        pdf.rect(10, pdf.get_y(), 190, 0.4, 'F')
    
    draw_table_header()

    # Table rows
    pdf.set_font('Arial', '', 7)
    for idx, t in enumerate(transactions):
        # Check for page break
        if pdf.get_y() > 252:
            pdf.add_page()
            pdf.section_title('Transaction Ledger (Continued)')
            draw_table_header()
            pdf.set_font('Arial', '', 7)

        # Alternating row colors
        if idx % 2 == 0:
            pdf.set_fill_color(252, 252, 255)
        else:
            pdf.set_fill_color(240, 242, 250)
        
        date_str = str(t['date'])[:10] if t['date'] else ''
        desc = str(t['description'] or '')[:32]
        cat = str(t['category'] or '')[:18]
        
        # Clean account name
        acc_raw = str(t['account_name'] or '')
        acc = acc_raw.split('] ')[-1] if '] ' in acc_raw else acc_raw
        acc = acc[:20]
        
        amount = t['amount'] or 0
        is_expense = t['type'] == 'expense'
        amount_str = f"{'- ' if is_expense else '+ '}Rp {amount:,.0f}"
        type_str = t['type'].upper() if t['type'] else ''

        # Draw row
        pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
        pdf.set_font('Arial', '', 6.5)
        pdf.cell(col_widths[0], 6, str(idx + 1), 0, 0, 'C', fill=True)
        pdf.set_text_color(*TreasuryPDF.BLACK)
        pdf.set_font('Arial', '', 7)
        pdf.cell(col_widths[1], 6, date_str, 0, 0, 'C', fill=True)
        pdf.cell(col_widths[2], 6, desc, 0, 0, 'L', fill=True)
        pdf.cell(col_widths[3], 6, cat, 0, 0, 'C', fill=True)
        pdf.cell(col_widths[4], 6, acc, 0, 0, 'C', fill=True)
        
        # Amount with color
        if is_expense:
            pdf.set_text_color(*TreasuryPDF.RED)
        else:
            pdf.set_text_color(*TreasuryPDF.GREEN)
        pdf.set_font('Arial', 'B', 7)
        pdf.cell(col_widths[5], 6, amount_str, 0, 0, 'R', fill=True)
        
        # Type badge
        pdf.set_font('Arial', 'B', 6)
        if is_expense:
            pdf.set_fill_color(255, 230, 230)
            pdf.set_text_color(*TreasuryPDF.RED)
        else:
            pdf.set_fill_color(230, 255, 235)
            pdf.set_text_color(*TreasuryPDF.GREEN)
        pdf.cell(col_widths[5 + 1] if len(col_widths) > 6 else 20, 6, type_str, 0, 0, 'C', fill=True)
        pdf.ln()
        pdf.set_font('Arial', '', 7)

        # Row bottom border
        pdf.set_draw_color(220, 220, 230)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())

    # === TOTAL ROW ===
    pdf.set_fill_color(*TreasuryPDF.NAVY)
    pdf.set_text_color(*TreasuryPDF.WHITE)
    pdf.set_font('Arial', 'B', 7)
    total_w = col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3] + col_widths[4]
    pdf.cell(total_w, 7, f'TOTAL  ({total_transactions} transactions)', 0, 0, 'R', fill=True)
    pdf.set_text_color(*TreasuryPDF.WHITE)
    net_display = f"Net: {'+ ' if net_flow >= 0 else '- '}Rp {abs(net_flow):,.0f}"
    pdf.cell(col_widths[5], 7, net_display, 0, 0, 'R', fill=True)
    pdf.cell(col_widths[6], 7, '', 0, 0, 'C', fill=True)
    pdf.ln()

    # === END NOTES ===
    pdf.ln(5)
    pdf.set_fill_color(*TreasuryPDF.LIGHT_GRAY)
    note_y = pdf.get_y()
    if note_y > 255:
        pdf.add_page()
        note_y = pdf.get_y()
    pdf.rect(10, note_y, 190, 16, 'F')
    pdf.set_fill_color(*TreasuryPDF.GOLD)
    pdf.rect(10, note_y, 190, 0.5, 'F')
    pdf.set_xy(14, note_y + 3)
    pdf.set_font('Arial', 'I', 7)
    pdf.set_text_color(*TreasuryPDF.DARK_GRAY)
    pdf.cell(0, 4, 'DISCLAIMER: This report was automatically generated by Treasury Flow Financial Platform.', 0, 1)
    pdf.set_x(14)
    pdf.cell(0, 4, 'All amounts are in Indonesian Rupiah (IDR). For questions or support, contact treasuryflow@gmail.com', 0, 1)

    # Output
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers.set('Content-Disposition', 'attachment', filename='Treasury_Flow_Financial_Report.pdf')
    response.headers.set('Content-Type', 'application/pdf')
    return response

@transactions_bp.route('/transaction/<int:id>/receipt')
@login_required
def download_receipt(id):
    user_id = session['user_id']
    db = get_db()
    tx = db.execute('''
        SELECT t.*, a.name as account_name 
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        WHERE t.id = ? AND t.user_id = ?
    ''', (id, user_id)).fetchone()
    
    if not tx:
        flash('Transaction not found', 'danger')
        return redirect(url_for('transactions.transactions_list'))

    user = db.execute('SELECT fullname, username FROM users WHERE id = ?', (user_id,)).fetchone()
    fullname = user['fullname'] if user and user['fullname'] else (user['username'] if user else 'User')

    NAVY = (17, 35, 126)
    GOLD = (180, 150, 60)
    DARK_GRAY = (100, 100, 110)
    ACCENT_BLUE = (60, 100, 200)
    BLACK = (30, 30, 30)
    GREEN = (16, 150, 72)
    RED = (200, 40, 40)
    LIGHT_GRAY = (245, 245, 248)

    pdf = FPDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    # Logo
    import os
    from flask import current_app
    logo_file = os.path.join(current_app.static_folder, 'img', 'logo_v4.png')

    # Top accent bar
    pdf.set_fill_color(*NAVY)
    pdf.rect(0, 0, 210, 3, 'F')
    pdf.set_fill_color(*GOLD)
    pdf.rect(0, 3, 210, 0.7, 'F')

    # Logo on the left
    if os.path.exists(logo_file):
        try:
            pdf.image(logo_file, 14, 6, 22)
        except Exception:
            pass

    # Header
    pdf.set_y(7)
    pdf.set_font('Arial', 'B', 22)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 10, 'TREASURY FLOW', 0, 1, 'C')
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(*DARK_GRAY)
    pdf.cell(0, 5, 'Personal Wealth Management & Financial Analytics', 0, 1, 'C')
    pdf.set_font('Arial', '', 7.5)
    pdf.set_text_color(*ACCENT_BLUE)
    pdf.cell(0, 4, 'treasuryflow@gmail.com  |  www.treasuryflow.web.id', 0, 1, 'C')
    pdf.set_fill_color(*NAVY)
    pdf.rect(10, 32, 190, 0.5, 'F')
    pdf.set_fill_color(*GOLD)
    pdf.rect(10, 32.5, 190, 0.3, 'F')

    # Title
    pdf.set_y(38)
    pdf.set_font('Arial', 'B', 16)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 10, 'TRANSACTION RECEIPT', 0, 1, 'C')
    pdf.ln(4)

    # Receipt info box
    is_expense = tx['type'] == 'expense'
    amount = tx['amount'] or 0
    acc_raw = str(tx['account_name'] or '')
    acc = acc_raw.split('] ')[-1] if '] ' in acc_raw else acc_raw

    box_y = pdf.get_y()
    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(20, box_y, 170, 60, 'F')

    fields = [
        ('Receipt No', f'#TF-{id:06d}'),
        ('Date', str(tx['date'])[:10]),
        ('Prepared For', fullname),
        ('Account', acc),
        ('Type', tx['type'].upper()),
        ('Category', str(tx['category'] or '-')),
        ('Description', str(tx['description'] or '-')[:50]),
    ]

    y_pos = box_y + 4
    for label, value in fields:
        pdf.set_xy(26, y_pos)
        pdf.set_font('Arial', '', 9)
        pdf.set_text_color(*DARK_GRAY)
        pdf.cell(35, 6, label, 0, 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.set_text_color(*BLACK)
        pdf.cell(120, 6, f':  {value}', 0, 0, 'L')
        y_pos += 7

    # Amount highlight
    pdf.set_y(box_y + 65)
    pdf.set_fill_color(*NAVY)
    pdf.rect(20, pdf.get_y(), 170, 18, 'F')
    pdf.set_xy(26, pdf.get_y() + 2)
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(40, 6, 'AMOUNT', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 18)
    color = RED if is_expense else GREEN
    pdf.set_text_color(*color)
    sign = '- ' if is_expense else '+ '
    pdf.cell(120, 12, f'{sign}Rp {amount:,.0f}', 0, 0, 'R')

    # Footer
    pdf.set_y(-40)
    pdf.set_fill_color(200, 200, 210)
    pdf.rect(10, pdf.get_y(), 190, 0.3, 'F')
    pdf.ln(4)
    pdf.set_font('Arial', 'I', 7)
    pdf.set_text_color(*DARK_GRAY)
    pdf.cell(0, 4, 'This receipt was automatically generated by Treasury Flow. No signature required.', 0, 1, 'C')
    pdf.set_font('Arial', 'B', 7)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 4, 'Powered by Fernando Capital', 0, 1, 'C')
    pdf.set_font('Arial', '', 7)
    pdf.set_text_color(*ACCENT_BLUE)
    pdf.cell(0, 4, f'Page {pdf.page_no()}/{{nb}}', 0, 0, 'C')

    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers.set('Content-Disposition', 'attachment', filename=f'Treasury_Flow_Receipt_{id}.pdf')
    response.headers.set('Content-Type', 'application/pdf')
    return response
